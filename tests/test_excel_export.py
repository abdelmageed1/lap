"""Regression test for the Excel (.xlsx) export used in Reports & Statistics, with an embedded
bar chart - complementing the existing plain-CSV export."""
import os

from openpyxl import load_workbook

from app.reports.excel_export import generate_excel_report


def test_generate_excel_report_writes_headers_rows_and_chart(tmp_path):
    path = str(tmp_path / "report.xlsx")
    generate_excel_report(
        path, "تقرير الأطباء المحولين", ["اسم الطبيب", "عدد الزيارات"],
        [["د. أحمد", 12], ["د. سارة", 8]],
        chart_label_col=0, chart_value_col=1, chart_title="أعلى الأطباء",
    )

    assert os.path.exists(path)
    wb = load_workbook(path)
    ws = wb.active
    assert ws["A1"].value == "تقرير الأطباء المحولين"
    assert ws["A3"].value == "اسم الطبيب"
    assert ws["B3"].value == "عدد الزيارات"
    assert ws["A4"].value == "د. أحمد"
    assert ws["B4"].value == 12
    assert len(ws._charts) == 1


def test_generate_excel_report_without_chart_columns_still_writes_data(tmp_path):
    path = str(tmp_path / "report_no_chart.xlsx")
    generate_excel_report(path, "تقرير بسيط", ["العمود"], [["قيمة"]])
    wb = load_workbook(path)
    ws = wb.active
    assert ws["A4"].value == "قيمة"
    assert len(ws._charts) == 0


def test_generate_excel_report_with_no_rows_does_not_crash(tmp_path):
    path = str(tmp_path / "report_empty.xlsx")
    generate_excel_report(path, "تقرير فارغ", ["العمود"], [], chart_label_col=0, chart_value_col=0)
    assert os.path.exists(path)
