"""Generates .xlsx report exports (with an embedded bar chart) for Reports & Statistics -
complementing the existing plain-CSV export with something presentable for management review."""
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)


def _write_sheet(ws, title: str, headers: list, rows: list,
                 chart_label_col: int = None, chart_value_col: int = None, chart_title: str = ""):
    ws.sheet_view.rightToLeft = True

    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))

    header_row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows] or [10])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    if rows and chart_label_col is not None and chart_value_col is not None:
        last_row = header_row + len(rows)
        chart = BarChart()
        chart.title = chart_title or headers[chart_value_col]
        chart.style = 10
        chart.y_axis.title = headers[chart_value_col]
        data = Reference(ws, min_col=chart_value_col + 1, min_row=header_row,
                          max_row=last_row, max_col=chart_value_col + 1)
        categories = Reference(ws, min_col=chart_label_col + 1, min_row=header_row + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width, chart.height = 20, 10
        ws.add_chart(chart, f"{get_column_letter(len(headers) + 2)}{header_row}")


def generate_excel_report(file_path: str, title: str, headers: list, rows: list,
                          chart_label_col: int = None, chart_value_col: int = None,
                          chart_title: str = "") -> None:
    """rows: list of row tuples/lists aligned with `headers`. If chart_label_col and
    chart_value_col are given (0-based indices into `headers`/each row), a bar chart is embedded
    using those two columns as categories/values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير"
    _write_sheet(ws, title, headers, rows, chart_label_col, chart_value_col, chart_title)
    wb.save(file_path)


def generate_periodic_report_excel(file_path: str, period_label: str, summary_rows: list,
                                   doctors_rows: list, staff_rows: list) -> None:
    """Multi-sheet workbook for the automated weekly/monthly report: an overall summary sheet,
    top referring doctors, and staff productivity - each for the elapsed period."""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "الملخص العام"
    _write_sheet(ws_summary, f"ملخص الأداء - {period_label}", ["البيان", "القيمة"], summary_rows)

    ws_doctors = wb.create_sheet("الأطباء المحوِّلون")
    _write_sheet(ws_doctors, f"أعلى الأطباء تحويلاً - {period_label}",
                 ["اسم الطبيب", "عدد الزيارات", "إجمالي الإيرادات"], doctors_rows,
                 chart_label_col=0, chart_value_col=2, chart_title="إيرادات الأطباء")

    ws_staff = wb.create_sheet("إنتاجية الموظفين")
    _write_sheet(ws_staff, f"إنتاجية الموظفين - {period_label}",
                 ["اسم الموظف", "الزيارات المسجَّلة", "التحصيلات", "النتائج المعالجة"], staff_rows,
                 chart_label_col=0, chart_value_col=1, chart_title="زيارات الموظفين")

    wb.save(file_path)
